from contextlib import asynccontextmanager
from datetime import date, datetime, time
from io import BytesIO
import re
import unicodedata
from typing import Literal
from zoneinfo import ZoneInfo

from fastapi import BackgroundTasks, Depends, FastAPI, File, Header, HTTPException, Query, UploadFile
from fastapi.responses import RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field, ValidationError
from openpyxl import load_workbook

from app.auth import (
    ensure_bootstrap_admin,
    get_current_user,
    hash_password,
    login_user,
    logout_session,
    require_roles,
)
from app.config import settings
from app.database import close_pool, connection, init_schema, start_pool
from app.email_service import notify_permission_event, send_test_email
from app.sheets_service import (
    create_authorization_url,
    disconnect as disconnect_sheets,
    finish_authorization,
    integration_status,
    sync_all,
)
from app.workflow import (
    active_boss,
    add_history,
    calculate_minutes,
    get_permission_for_user,
    max_business_date,
    minutes_between,
)


@asynccontextmanager
async def lifespan(app: FastAPI):
    start_pool()
    init_schema()
    ensure_bootstrap_admin()
    yield
    close_pool()


app = FastAPI(title="ERSeP · Permisos de Salida API", version="0.7.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=list(settings.frontend_origins),
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)


class LoginIn(BaseModel):
    usuario: str = Field(min_length=1, max_length=80)
    clave: str = Field(min_length=1, max_length=200)


class PermissionIn(BaseModel):
    tipo: Literal["OFICIAL", "PARTICULAR"]
    fecha_salida: date
    lugar_destino: str | None = Field(default=None, max_length=300)
    hora_salida: str
    hora_regreso: str | None = None
    sin_regreso: bool = False
    minutos_declarados: int | None = Field(default=None, ge=0, le=24 * 60)
    justificacion_minutos: str | None = Field(default=None, max_length=1000)
    fecha_devolucion: date | None = None
    devolucion_hora_desde: str | None = None
    devolucion_hora_hasta: str | None = None
    compensacion_modo: Literal["DEVOLVER_HORAS", "HORAS_EXTRAS_PREVIAS"] = "DEVOLVER_HORAS"
    horas_extra_fecha: date | None = None
    horas_extra_desde: str | None = None
    horas_extra_hasta: str | None = None
    justificacion_fuera_plazo: str | None = Field(default=None, max_length=1000)
    observaciones: str | None = Field(default=None, max_length=1000)


class DecisionIn(BaseModel):
    observacion: str | None = Field(default=None, max_length=1000)


class AdminUserIn(BaseModel):
    username: str = Field(min_length=3, max_length=80, pattern=r"^[A-Za-z0-9._-]+$")
    password: str | None = Field(default=None, min_length=6, max_length=200)
    email: EmailStr
    nombre: str = Field(min_length=1, max_length=120)
    apellido: str = Field(min_length=1, max_length=120)
    legajo: str = Field(min_length=1, max_length=40)
    dni: str | None = Field(default=None, max_length=30)
    area: str | None = Field(default=None, max_length=180)  # compatibilidad heredada
    oficina_id: int | None = None
    jornada_desde: str = "08:00"
    jornada_hasta: str = "14:00"
    roles: list[Literal["AGENTE", "JEFE", "RRHH", "ADMIN"]] = ["AGENTE"]
    jefe_email: EmailStr | None = None  # compatibilidad heredada, ya no se usa en la UI V6


class AdminOfficeIn(BaseModel):
    id: int | None = None
    nombre: str = Field(min_length=1, max_length=180)
    jefe_id: int | None = None
    activo: bool = True


class AdminUserStatusIn(BaseModel):
    activo: bool


class CleanupIn(BaseModel):
    confirmacion: str = Field(min_length=1, max_length=80)


def _parse_time(value: str | None):
    if value is None or value == "":
        return None
    try:
        hh, mm = value.split(":")[:2]
        parsed = time(hour=int(hh), minute=int(mm))
        return parsed
    except Exception:
        raise HTTPException(status_code=422, detail=f"Hora inválida: {value}")


def _serialize_permission(row: dict):
    row = dict(row)
    for key in (
        "hora_salida", "hora_regreso", "jornada_desde", "jornada_hasta",
        "reposicion_hora_desde", "reposicion_hora_hasta",
        "hora_desde_horas_extra", "hora_hasta_horas_extra",
    ):
        if row.get(key) is not None:
            row[key] = str(row[key])[:5]

    declared = row.get("minutos_declarados")
    if declared is None:
        declared = row.get("minutos_autorizados")
    mode = row.get("reposicion_modalidad") or row.get("modalidad_compensacion") or "DEVOLVER_HORAS"
    proposed = row.get("minutos_horas_extra") if mode == "HORAS_EXTRAS_PREVIAS" else row.get("reposicion_minutos_tramo")
    if proposed is None and mode == "DEVOLVER_HORAS":
        a = row.get("reposicion_hora_desde")
        b = row.get("reposicion_hora_hasta")
        if a and b:
            try:
                ah, am = map(int, str(a).split(":")[:2]); bh, bm = map(int, str(b).split(":")[:2])
                proposed = (bh * 60 + bm) - (ah * 60 + am)
            except Exception:
                proposed = None

    risks = []
    if row.get("fuera_plazo_reglamentario"):
        risks.append({"nivel": "CRITICO", "codigo": "FUERA_PLAZO", "mensaje": "La devolución propuesta supera el plazo reglamentario sugerido."})
    if declared is not None and proposed is not None and int(proposed) < int(declared):
        label = "horas extra informadas" if mode == "HORAS_EXTRAS_PREVIAS" else "tramo de devolución"
        risks.append({"nivel": "CRITICO", "codigo": "COMPENSACION_INSUFICIENTE", "mensaje": f"El {label} cubre menos tiempo que la salida declarada ({proposed} min vs. {declared} min)."})
    calculated = row.get("minutos_calculados")
    if declared is not None and calculated is not None and int(declared) != int(calculated):
        risks.append({"nivel": "ATENCION", "codigo": "TIEMPO_DIFERENTE", "mensaje": "El tiempo declarado por el agente difiere del cálculo automático."})
    row["riesgos"] = risks
    row["riesgo_critico"] = any(x["nivel"] == "CRITICO" for x in risks)
    return row


def _serialize_user(row: dict):
    row = dict(row)
    row.pop("password_hash", None)
    row.pop("google_sub", None)
    for key in ("jornada_desde", "jornada_hasta"):
        if row.get(key) is not None:
            row[key] = str(row[key])[:5]
    return row


@app.get("/api/health")
def health():
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 ok")
            db_ok = cur.fetchone()["ok"] == 1
    try:
        sheets = integration_status()
    except Exception:
        sheets = {"configured": False, "authorized": False}
    return {
        "status": "ok",
        "database": db_ok,
        "auth": "local",
        "bootstrap_configured": bool(settings.bootstrap_admin_password),
        "sheets_configured": bool(sheets.get("configured")),
        "sheets_authorized": bool(sheets.get("authorized")),
    }


@app.post("/api/auth/login")
def auth_login(payload: LoginIn):
    result = login_user(payload.usuario, payload.clave)
    result["user"] = _serialize_user(result["user"])
    return result


@app.post("/api/auth/logout")
def auth_logout(authorization: str | None = Header(default=None)):
    logout_session(authorization)
    return {"status": "ok"}


@app.get("/api/auth/me")
def auth_me(user: dict = Depends(get_current_user)):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT o.id oficina_id,o.nombre oficina,
                       trim(concat(j.nombre,' ',j.apellido)) jefe_nombre,j.email jefe_email
                FROM usuarios u
                LEFT JOIN oficinas o ON o.id=u.oficina_id
                LEFT JOIN usuarios j ON j.id=o.jefe_id
                WHERE u.id=%s
            """, (user["id"],))
            org = cur.fetchone() or {}
    return _serialize_user({
        "id": user["id"],
        "email": user["email"],
        "nombre": user["nombre"],
        "apellido": user["apellido"],
        "dni": user.get("dni"),
        "legajo": user.get("legajo"),
        "area": org.get("oficina") or user.get("area"),
        "oficina_id": org.get("oficina_id"),
        "oficina": org.get("oficina") or user.get("area"),
        "jefe_nombre": org.get("jefe_nombre"),
        "jefe_email": org.get("jefe_email"),
        "jornada_desde": user.get("jornada_desde"),
        "jornada_hasta": user.get("jornada_hasta"),
        "roles": user["roles"],
    })


@app.get("/api/reglas/plazo-devolucion")
def return_deadline(fecha_salida: date = Query(...), user: dict = Depends(require_roles("AGENTE"))):
    with connection() as conn:
        with conn.cursor() as cur:
            limit_date = max_business_date(cur, fecha_salida, 7)
    return {
        "fecha_salida": fecha_salida,
        "fecha_limite": limit_date,
        "dias_habiles": 7,
        "mensaje": "Por reglamento, la devolución debería realizarse dentro de los próximos 7 días hábiles.",
    }


@app.post("/api/permisos")
def create_permission(payload: PermissionIn, bg: BackgroundTasks, user: dict = Depends(require_roles("AGENTE"))):
    if payload.fecha_salida < datetime.now(ZoneInfo("America/Argentina/Cordoba")).date():
        raise HTTPException(status_code=422, detail="La fecha de salida no puede ser anterior al día de hoy.")
    if payload.tipo == "OFICIAL" and not (payload.lugar_destino or "").strip():
        raise HTTPException(status_code=422, detail="Las salidas oficiales requieren lugar de destino.")

    start = _parse_time(payload.hora_salida)
    end = _parse_time(payload.hora_regreso)
    workday_start = user.get("jornada_desde") or time(8, 0)
    workday_end = user.get("jornada_hasta") or time(14, 0)
    calculated = calculate_minutes(start, end, payload.sin_regreso, workday_end)
    declared = calculated if payload.minutos_declarados is None else payload.minutos_declarados

    if declared != calculated and not (payload.justificacion_minutos or "").strip():
        raise HTTPException(
            status_code=422,
            detail="El tiempo de salida declarado difiere del cálculo automático. Debe explicar el motivo de la diferencia.",
        )

    mode = payload.compensacion_modo if payload.tipo == "PARTICULAR" else None
    return_from = _parse_time(payload.devolucion_hora_desde)
    return_to = _parse_time(payload.devolucion_hora_hasta)
    extra_from = _parse_time(payload.horas_extra_desde)
    extra_to = _parse_time(payload.horas_extra_hasta)

    with connection() as conn:
        with conn.cursor() as cur:
            limit_date = None
            outside = False
            return_minutes = None
            extra_minutes = None
            return_date = None
            extra_date = None

            if payload.tipo == "PARTICULAR":
                if mode == "DEVOLVER_HORAS":
                    if not payload.fecha_devolucion:
                        raise HTTPException(status_code=422, detail="Debe indicar la fecha en la que devolverá las horas.")
                    if payload.fecha_devolucion < payload.fecha_salida:
                        raise HTTPException(status_code=422, detail="La fecha de devolución no puede ser anterior a la salida.")
                    if return_from is None or return_to is None:
                        raise HTTPException(status_code=422, detail="Debe indicar el horario completo en el que devolverá las horas.")
                    return_minutes = minutes_between(return_from, return_to)
                    return_date = payload.fecha_devolucion
                    limit_date = max_business_date(cur, payload.fecha_salida, 7)
                    outside = payload.fecha_devolucion > limit_date
                    if outside and not (payload.justificacion_fuera_plazo or "").strip():
                        raise HTTPException(
                            status_code=422,
                            detail=(
                                f"La fecha seleccionada supera el plazo reglamentario sugerido ({limit_date.strftime('%d/%m/%Y')}). "
                                "Puede continuar, pero debe indicar una observación para consideración de RR.HH."
                            ),
                        )
                elif mode == "HORAS_EXTRAS_PREVIAS":
                    if not payload.horas_extra_fecha or extra_from is None or extra_to is None:
                        raise HTTPException(
                            status_code=422,
                            detail="Para usar horas extras previas debe indicar el día y el horario exacto en que fueron realizadas.",
                        )
                    if payload.horas_extra_fecha >= payload.fecha_salida:
                        raise HTTPException(
                            status_code=422,
                            detail="Las horas extra utilizadas deben haber sido realizadas antes de la fecha de salida.",
                        )
                    extra_minutes = minutes_between(extra_from, extra_to)
                    extra_date = payload.horas_extra_fecha
                else:
                    raise HTTPException(status_code=422, detail="Modalidad de compensación inválida.")

            cur.execute("""
                INSERT INTO permisos_salida (
                  agente_id,oficina_id,tipo,fecha_salida,lugar_destino,hora_salida,hora_regreso,sin_regreso,
                  jornada_desde,jornada_hasta,minutos_calculados,minutos_declarados,minutos_autorizados,
                  justificacion_minutos,fecha_devolucion,fecha_limite_devolucion,fuera_plazo_reglamentario,
                  justificacion_fuera_plazo,modalidad_compensacion,observaciones,estado
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'BORRADOR') RETURNING *
            """, (
                user["id"], user.get("oficina_id"), payload.tipo, payload.fecha_salida,
                (payload.lugar_destino or "").strip() or None,
                start, end, payload.sin_regreso, workday_start, workday_end, calculated, declared, declared,
                (payload.justificacion_minutos or "").strip() or None, return_date, limit_date, outside,
                (payload.justificacion_fuera_plazo or "").strip() or None, mode, payload.observaciones,
            ))
            p = cur.fetchone()
            number = f"PS-{payload.fecha_salida.year}-{p['id']:06d}"
            cur.execute("UPDATE permisos_salida SET numero_permiso=%s WHERE id=%s", (number, p["id"]))

            details = [
                f"Jornada {str(workday_start)[:5]}–{str(workday_end)[:5]}",
                f"Cálculo automático: {calculated} min",
                f"Tiempo de salida declarado: {declared} min",
            ]
            if declared != calculated:
                details.append(f"Justificación: {(payload.justificacion_minutos or '').strip()}")
            if payload.tipo == "OFICIAL":
                details.append("Salida oficial: no corresponde devolución ni compensación de horas; requiere aprobación final de RR.HH.")
            if mode == "DEVOLVER_HORAS":
                details.append(f"Compensación: devolución el {return_date.strftime('%d/%m/%Y')} de {str(return_from)[:5]} a {str(return_to)[:5]} ({return_minutes} min)")
            elif mode == "HORAS_EXTRAS_PREVIAS":
                details.append(f"Compensación: horas extra previas del {extra_date.strftime('%d/%m/%Y')} de {str(extra_from)[:5]} a {str(extra_to)[:5]} ({extra_minutes} min)")
            if outside:
                details.append(f"Devolución fuera del plazo sugerido ({limit_date.strftime('%d/%m/%Y')}): {(payload.justificacion_fuera_plazo or '').strip()}")
            add_history(cur, p["id"], user["id"], "SOLICITUD_CREADA", None, "BORRADOR", " · ".join(details))

            if payload.tipo == "PARTICULAR":
                cur.execute("""
                    INSERT INTO reposiciones (
                        permiso_id,fecha_prevista,hora_desde_prevista,hora_hasta_prevista,minutos_a_reponer,
                        modalidad,fecha_horas_extra,hora_desde_horas_extra,hora_hasta_horas_extra,minutos_horas_extra
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    p["id"], return_date, return_from, return_to, declared,
                    mode, extra_date, extra_from, extra_to, extra_minutes,
                ))
                compensated = extra_minutes if mode == "HORAS_EXTRAS_PREVIAS" else return_minutes
                if compensated != declared:
                    add_history(
                        cur, p["id"], user["id"], "COMPENSACION_DIFERENTE", "BORRADOR", "BORRADOR",
                        f"La compensación informada equivale a {compensated} min y el tiempo de salida declarado es {declared} min.",
                    )
            conn.commit()
            p["numero_permiso"] = number
            result = _serialize_permission(p)

    if settings.sheets_enabled:
        bg.add_task(sync_all)
    return result


@app.post("/api/permisos/{permission_id}/enviar")
def send_permission(permission_id: int, bg: BackgroundTasks, user: dict = Depends(require_roles("AGENTE"))):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM permisos_salida WHERE id=%s FOR UPDATE", (permission_id,))
            p = cur.fetchone()
            if not p or p["agente_id"] != user["id"]:
                raise HTTPException(status_code=404, detail="Permiso inexistente.")
            if p["estado"] != "BORRADOR":
                raise HTTPException(status_code=409, detail="Solo un borrador puede enviarse a autorización.")
            boss = active_boss(cur, user["id"], p["fecha_salida"])
            if not boss:
                raise HTTPException(status_code=409, detail="No tenés una Oficina con jefatura activa configurada para esa fecha. Solicitá a Administración o RR.HH. que revise tu Oficina.")
            cur.execute("UPDATE permisos_salida SET estado='PENDIENTE_JEFE',jefe_asignado_id=%s,updated_at=NOW() WHERE id=%s", (boss["id"], permission_id))
            add_history(cur, permission_id, user["id"], "ENVIADO_A_JEFE", "BORRADOR", "PENDIENTE_JEFE", f"Jefatura asignada: {boss['nombre']} {boss['apellido']}")
            conn.commit()
    if settings.sheets_enabled:
        bg.add_task(sync_all)
    bg.add_task(notify_permission_event, permission_id, "SOLICITUD_ENVIADA", None)
    return {"status": "ok", "message": "Solicitud enviada a autorización."}


@app.get("/api/permisos/mios")
def my_permissions(user: dict = Depends(require_roles("AGENTE"))):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.*,trim(concat(u.nombre,' ',u.apellido)) agente_nombre,u.legajo,
                       COALESCE(op.nombre,ou.nombre,u.area) oficina,
                       trim(concat(j.nombre,' ',j.apellido)) jefe_nombre,
                       r.fecha_prevista reposicion_fecha_prevista,
                       r.hora_desde_prevista reposicion_hora_desde,r.hora_hasta_prevista reposicion_hora_hasta,
                       CASE WHEN r.hora_desde_prevista IS NOT NULL AND r.hora_hasta_prevista IS NOT NULL
                            THEN EXTRACT(EPOCH FROM (r.hora_hasta_prevista-r.hora_desde_prevista))/60 END reposicion_minutos_tramo,
                       r.minutos_a_reponer reposicion_minutos,r.modalidad reposicion_modalidad,
                       r.fecha_horas_extra,r.hora_desde_horas_extra,r.hora_hasta_horas_extra,r.minutos_horas_extra
                FROM permisos_salida p
                JOIN usuarios u ON u.id=p.agente_id
                LEFT JOIN oficinas op ON op.id=p.oficina_id
                LEFT JOIN oficinas ou ON ou.id=u.oficina_id
                LEFT JOIN usuarios j ON j.id=p.jefe_asignado_id
                LEFT JOIN reposiciones r ON r.permiso_id=p.id
                WHERE p.agente_id=%s ORDER BY p.fecha_salida DESC,p.id DESC
            """, (user["id"],))
            rows = [_serialize_permission(r) for r in cur.fetchall()]
    return {"items": rows}


@app.get("/api/permisos/{permission_id}")
def permission_detail(permission_id: int, user: dict = Depends(get_current_user)):
    return _serialize_permission(get_permission_for_user(permission_id, user))


def _jefatura_rows(user_id: int, estado: str | None = None, tipo: str | None = None,
                    agente_id: int | None = None, fecha_desde: date | None = None,
                    fecha_hasta: date | None = None, limit: int = 2000):
    clauses = ["(p.jefe_asignado_id=%s OR EXISTS (SELECT 1 FROM oficinas oj WHERE oj.id=p.oficina_id AND oj.jefe_id=%s))"]
    params: list = [user_id, user_id]
    if estado:
        clauses.append("p.estado=%s"); params.append(estado)
    if tipo:
        clauses.append("p.tipo=%s"); params.append(tipo)
    if agente_id:
        clauses.append("p.agente_id=%s"); params.append(agente_id)
    if fecha_desde:
        clauses.append("p.fecha_salida >= %s"); params.append(fecha_desde)
    if fecha_hasta:
        clauses.append("p.fecha_salida <= %s"); params.append(fecha_hasta)
    params.append(limit)
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT p.*,trim(concat(a.nombre,' ',a.apellido)) agente_nombre,a.legajo,
                       COALESCE(op.nombre,ou.nombre,a.area) oficina,
                       trim(concat(j.nombre,' ',j.apellido)) jefe_nombre,
                       r.fecha_prevista reposicion_fecha_prevista,
                       r.hora_desde_prevista reposicion_hora_desde,r.hora_hasta_prevista reposicion_hora_hasta,
                       CASE WHEN r.hora_desde_prevista IS NOT NULL AND r.hora_hasta_prevista IS NOT NULL
                            THEN EXTRACT(EPOCH FROM (r.hora_hasta_prevista-r.hora_desde_prevista))/60 END reposicion_minutos_tramo,
                       r.minutos_a_reponer reposicion_minutos,r.modalidad reposicion_modalidad,
                       r.fecha_horas_extra,r.hora_desde_horas_extra,r.hora_hasta_horas_extra,r.minutos_horas_extra,
                       aj.decision decision_jefatura,aj.fecha_hora decision_jefatura_fecha
                FROM permisos_salida p
                JOIN usuarios a ON a.id=p.agente_id
                LEFT JOIN oficinas op ON op.id=p.oficina_id
                LEFT JOIN oficinas ou ON ou.id=a.oficina_id
                LEFT JOIN usuarios j ON j.id=p.jefe_asignado_id
                LEFT JOIN reposiciones r ON r.permiso_id=p.id
                LEFT JOIN LATERAL (
                    SELECT ap.decision,ap.fecha_hora
                    FROM aprobaciones ap
                    WHERE ap.permiso_id=p.id AND ap.tipo_aprobacion='JEFE'
                    ORDER BY ap.fecha_hora DESC,ap.id DESC LIMIT 1
                ) aj ON TRUE
                WHERE {' AND '.join(clauses)}
                ORDER BY p.fecha_salida DESC,p.id DESC LIMIT %s
            """, params)
            return [_serialize_permission(r) for r in cur.fetchall()]


@app.get("/api/jefatura/agentes")
def boss_agents(user: dict = Depends(require_roles("JEFE"))):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT u.id,u.nombre,u.apellido,u.legajo,u.dni,u.email,
                       to_char(u.jornada_desde,'HH24:MI') jornada_desde,to_char(u.jornada_hasta,'HH24:MI') jornada_hasta,o.id oficina_id,o.nombre oficina
                FROM usuarios u
                JOIN oficinas o ON o.id=u.oficina_id
                WHERE u.activo=TRUE AND o.activo=TRUE AND o.jefe_id=%s
                  AND EXISTS (
                    SELECT 1 FROM usuario_roles ur JOIN roles r ON r.id=ur.rol_id
                    WHERE ur.usuario_id=u.id AND r.codigo='AGENTE'
                  )
                ORDER BY u.apellido,u.nombre,u.legajo
            """, (user["id"],))
            return {"items": cur.fetchall()}


@app.get("/api/jefatura/pendientes")
def boss_queue(user: dict = Depends(require_roles("JEFE"))):
    return {"items": _jefatura_rows(user["id"], estado="PENDIENTE_JEFE")}


@app.get("/api/jefatura/permisos")
def boss_permissions(
    estado: str | None = Query(default=None),
    tipo: str | None = Query(default=None),
    agente_id: int | None = Query(default=None),
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    user: dict = Depends(require_roles("JEFE")),
):
    return {"items": _jefatura_rows(user["id"], estado, tipo, agente_id, fecha_desde, fecha_hasta)}


@app.get("/api/jefatura/dashboard")
def boss_dashboard(
    estado: str | None = Query(default=None),
    tipo: str | None = Query(default=None),
    agente_id: int | None = Query(default=None),
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    user: dict = Depends(require_roles("JEFE")),
):
    rows = _jefatura_rows(user["id"], estado, tipo, agente_id, fecha_desde, fecha_hasta)
    total = len(rows)
    by_agent = {}
    by_hour = {}
    by_office = {}
    for r in rows:
        by_agent[r.get("agente_nombre") or "Sin identificar"] = by_agent.get(r.get("agente_nombre") or "Sin identificar", 0) + 1
        hh = str(r.get("hora_salida") or "—")[:2]
        hour_label = f"{hh}:00" if hh.isdigit() else "Sin horario"
        by_hour[hour_label] = by_hour.get(hour_label, 0) + 1
        office = r.get("oficina") or "Sin Oficina"
        by_office[office] = by_office.get(office, 0) + 1
    top = lambda d, n=8: [{"label": k, "valor": v} for k, v in sorted(d.items(), key=lambda x: (-x[1], x[0]))[:n]]
    return {
        "total": total,
        "pendientes": sum(1 for r in rows if r.get("estado") == "PENDIENTE_JEFE"),
        "autorizados": sum(1 for r in rows if r.get("decision_jefatura") == "APROBADO"),
        "rechazados": sum(1 for r in rows if r.get("decision_jefatura") == "RECHAZADO"),
        "criticos": sum(1 for r in rows if r.get("riesgo_critico")),
        "por_agente": top(by_agent),
        "por_hora": top(by_hour, 12),
        "por_oficina": top(by_office),
    }


def _can_boss_act(cur, permission: dict, user: dict) -> bool:
    if "ADMIN" in user.get("roles", []):
        return True
    if permission.get("jefe_asignado_id") == user.get("id"):
        return True
    office_id = permission.get("oficina_id")
    if not office_id:
        return False
    cur.execute("SELECT 1 FROM oficinas WHERE id=%s AND jefe_id=%s AND activo=TRUE", (office_id, user["id"]))
    return cur.fetchone() is not None


@app.post("/api/permisos/{permission_id}/autorizar")
def authorize(permission_id: int, payload: DecisionIn, bg: BackgroundTasks, user: dict = Depends(require_roles("JEFE"))):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM permisos_salida WHERE id=%s FOR UPDATE", (permission_id,))
            p = cur.fetchone()
            if not p:
                raise HTTPException(status_code=404, detail="Permiso inexistente.")
            if not _can_boss_act(cur, p, user):
                raise HTTPException(status_code=403, detail="No sos la jefatura asignada ni la jefatura actual de esta Oficina.")
            if p["estado"] != "PENDIENTE_JEFE":
                raise HTTPException(status_code=409, detail="La solicitud ya no está pendiente de autorización.")
            cur.execute("INSERT INTO aprobaciones (permiso_id,usuario_id,tipo_aprobacion,decision,observacion) VALUES (%s,%s,'JEFE','APROBADO',%s)", (permission_id, user["id"], payload.observacion))
            cur.execute("UPDATE permisos_salida SET estado='PENDIENTE_RRHH',updated_at=NOW() WHERE id=%s", (permission_id,))
            add_history(cur, permission_id, user["id"], "AUTORIZADO_JEFE", "PENDIENTE_JEFE", "PENDIENTE_RRHH", payload.observacion)
            conn.commit()
    if settings.sheets_enabled:
        bg.add_task(sync_all)
    bg.add_task(notify_permission_event, permission_id, "JEFATURA_APROBADO", payload.observacion)
    return {"status": "ok"}


@app.post("/api/permisos/{permission_id}/rechazar")
def reject(permission_id: int, payload: DecisionIn, bg: BackgroundTasks, user: dict = Depends(require_roles("JEFE"))):
    reason = (payload.observacion or "").strip()
    if not reason:
        raise HTTPException(status_code=422, detail="Debe indicar el motivo del rechazo.")
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM permisos_salida WHERE id=%s FOR UPDATE", (permission_id,))
            p = cur.fetchone()
            if not p:
                raise HTTPException(status_code=404, detail="Permiso inexistente.")
            if not _can_boss_act(cur, p, user):
                raise HTTPException(status_code=403, detail="No sos la jefatura asignada ni la jefatura actual de esta Oficina.")
            if p["estado"] != "PENDIENTE_JEFE":
                raise HTTPException(status_code=409, detail="La solicitud ya no está pendiente.")
            cur.execute("INSERT INTO aprobaciones (permiso_id,usuario_id,tipo_aprobacion,decision,observacion) VALUES (%s,%s,'JEFE','RECHAZADO',%s)", (permission_id, user["id"], reason))
            cur.execute("UPDATE permisos_salida SET estado='RECHAZADO_JEFE',updated_at=NOW() WHERE id=%s", (permission_id,))
            add_history(cur, permission_id, user["id"], "RECHAZADO_JEFE", "PENDIENTE_JEFE", "RECHAZADO_JEFE", reason)
            conn.commit()
    if settings.sheets_enabled:
        bg.add_task(sync_all)
    bg.add_task(notify_permission_event, permission_id, "JEFATURA_RECHAZADO", reason)
    return {"status": "ok"}


def _rrhh_rows(
    estado: str | None = None,
    tipo: str | None = None,
    oficina_id: int | None = None,
    agente_id: int | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    limit: int = 3000,
):
    clauses = []
    params: list = []
    if estado:
        clauses.append("p.estado=%s"); params.append(estado)
    if tipo:
        clauses.append("p.tipo=%s"); params.append(tipo)
    if oficina_id:
        clauses.append("COALESCE(p.oficina_id,a.oficina_id)=%s"); params.append(oficina_id)
    if agente_id:
        clauses.append("p.agente_id=%s"); params.append(agente_id)
    if fecha_desde:
        clauses.append("p.fecha_salida >= %s"); params.append(fecha_desde)
    if fecha_hasta:
        clauses.append("p.fecha_salida <= %s"); params.append(fecha_hasta)
    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    params.append(limit)
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT p.*,trim(concat(a.nombre,' ',a.apellido)) agente_nombre,a.legajo,a.email agente_email,
                       COALESCE(op.nombre,ou.nombre,a.area) oficina,
                       trim(concat(j.nombre,' ',j.apellido)) jefe_nombre,
                       r.fecha_prevista reposicion_fecha_prevista,
                       r.hora_desde_prevista reposicion_hora_desde,r.hora_hasta_prevista reposicion_hora_hasta,
                       CASE WHEN r.hora_desde_prevista IS NOT NULL AND r.hora_hasta_prevista IS NOT NULL
                            THEN EXTRACT(EPOCH FROM (r.hora_hasta_prevista-r.hora_desde_prevista))/60 END reposicion_minutos_tramo,
                       r.minutos_a_reponer reposicion_minutos,r.modalidad reposicion_modalidad,
                       r.fecha_horas_extra,r.hora_desde_horas_extra,r.hora_hasta_horas_extra,r.minutos_horas_extra
                FROM permisos_salida p
                JOIN usuarios a ON a.id=p.agente_id
                LEFT JOIN oficinas op ON op.id=p.oficina_id
                LEFT JOIN oficinas ou ON ou.id=a.oficina_id
                LEFT JOIN usuarios j ON j.id=p.jefe_asignado_id
                LEFT JOIN reposiciones r ON r.permiso_id=p.id
                {where} ORDER BY p.fecha_salida DESC,p.id DESC LIMIT %s
            """, params)
            return [_serialize_permission(r) for r in cur.fetchall()]


@app.get("/api/rrhh/agentes")
def rrhh_agents(
    oficina_id: int | None = Query(default=None),
    user: dict = Depends(require_roles("RRHH")),
):
    params = []
    where = ["u.activo=TRUE"]
    if oficina_id:
        where.append("u.oficina_id=%s"); params.append(oficina_id)
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT u.id,u.nombre,u.apellido,u.legajo,u.dni,u.email,
                       to_char(u.jornada_desde,'HH24:MI') jornada_desde,to_char(u.jornada_hasta,'HH24:MI') jornada_hasta,u.oficina_id,o.nombre oficina
                FROM usuarios u
                LEFT JOIN oficinas o ON o.id=u.oficina_id
                WHERE {' AND '.join(where)}
                  AND EXISTS (
                    SELECT 1 FROM usuario_roles ur JOIN roles r ON r.id=ur.rol_id
                    WHERE ur.usuario_id=u.id AND r.codigo='AGENTE'
                  )
                ORDER BY o.nombre NULLS LAST,u.apellido,u.nombre
            """, params)
            return {"items": cur.fetchall()}


@app.get("/api/rrhh/permisos")
def rrhh_permissions(
    estado: str | None = Query(default=None),
    tipo: str | None = Query(default=None),
    oficina_id: int | None = Query(default=None),
    agente_id: int | None = Query(default=None),
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    user: dict = Depends(require_roles("RRHH")),
):
    return {"items": _rrhh_rows(estado, tipo, oficina_id, agente_id, fecha_desde, fecha_hasta)}


@app.get("/api/rrhh/dashboard")
def rrhh_dashboard(
    estado: str | None = Query(default=None),
    tipo: str | None = Query(default=None),
    oficina_id: int | None = Query(default=None),
    agente_id: int | None = Query(default=None),
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    user: dict = Depends(require_roles("RRHH")),
):
    rows = _rrhh_rows(estado, tipo, oficina_id, agente_id, fecha_desde, fecha_hasta)
    by_office = {}
    by_agent = {}
    by_hour = {}
    by_type = {}
    declared_minutes = 0
    for r in rows:
        office = r.get("oficina") or "Sin Oficina"
        agent_name = r.get("agente_nombre") or "Sin identificar"
        hh = str(r.get("hora_salida") or "—")[:2]
        hour_label = f"{hh}:00" if hh.isdigit() else "Sin horario"
        by_office[office] = by_office.get(office, 0) + 1
        by_agent[agent_name] = by_agent.get(agent_name, 0) + 1
        by_hour[hour_label] = by_hour.get(hour_label, 0) + 1
        by_type[r.get("tipo") or "—"] = by_type.get(r.get("tipo") or "—", 0) + 1
        if r.get("tipo") == "PARTICULAR":
            declared_minutes += int(r.get("minutos_declarados") or 0)
    top = lambda d, n=10: [{"label": k, "valor": v} for k, v in sorted(d.items(), key=lambda x: (-x[1], x[0]))[:n]]
    return {
        "total": len(rows),
        "pendientes_rrhh": sum(1 for r in rows if r.get("estado") == "PENDIENTE_RRHH"),
        "particulares": sum(1 for r in rows if r.get("tipo") == "PARTICULAR"),
        "minutos_particulares": declared_minutes,
        "criticos": sum(1 for r in rows if r.get("riesgo_critico")),
        "por_oficina": top(by_office),
        "agentes_recurrentes": top(by_agent),
        "por_hora": top(by_hour, 12),
        "por_tipo": top(by_type, 4),
    }


@app.post("/api/permisos/{permission_id}/verificar-rrhh")
def verify_rrhh(permission_id: int, payload: DecisionIn, bg: BackgroundTasks, user: dict = Depends(require_roles("RRHH"))):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM permisos_salida WHERE id=%s FOR UPDATE", (permission_id,))
            p = cur.fetchone()
            if not p:
                raise HTTPException(status_code=404, detail="Permiso inexistente.")
            if p["estado"] != "PENDIENTE_RRHH":
                raise HTTPException(status_code=409, detail="El permiso no está pendiente de RR.HH.")
            cur.execute("INSERT INTO aprobaciones (permiso_id,usuario_id,tipo_aprobacion,decision,observacion) VALUES (%s,%s,'RRHH','VERIFICADO',%s)", (permission_id, user["id"], payload.observacion))
            cur.execute("UPDATE permisos_salida SET estado='VERIFICADO_RRHH',updated_at=NOW() WHERE id=%s", (permission_id,))
            add_history(cur, permission_id, user["id"], "VERIFICADO_RRHH", "PENDIENTE_RRHH", "VERIFICADO_RRHH", payload.observacion)
            conn.commit()
    if settings.sheets_enabled:
        bg.add_task(sync_all)
    bg.add_task(notify_permission_event, permission_id, "RRHH_APROBADO", payload.observacion)
    return {"status": "ok"}


@app.post("/api/permisos/{permission_id}/rechazar-rrhh")
def reject_rrhh(permission_id: int, payload: DecisionIn, bg: BackgroundTasks, user: dict = Depends(require_roles("RRHH"))):
    reason = (payload.observacion or "").strip()
    if not reason:
        raise HTTPException(status_code=422, detail="RR.HH. debe indicar obligatoriamente el motivo del rechazo.")
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM permisos_salida WHERE id=%s FOR UPDATE", (permission_id,))
            p = cur.fetchone()
            if not p:
                raise HTTPException(status_code=404, detail="Permiso inexistente.")
            if p["estado"] != "PENDIENTE_RRHH":
                raise HTTPException(status_code=409, detail="El permiso no está pendiente de RR.HH.")
            cur.execute("INSERT INTO aprobaciones (permiso_id,usuario_id,tipo_aprobacion,decision,observacion) VALUES (%s,%s,'RRHH','RECHAZADO',%s)", (permission_id, user["id"], reason))
            cur.execute("UPDATE permisos_salida SET estado='RECHAZADO_RRHH',updated_at=NOW() WHERE id=%s", (permission_id,))
            add_history(cur, permission_id, user["id"], "RECHAZADO_RRHH", "PENDIENTE_RRHH", "RECHAZADO_RRHH", reason)
            conn.commit()
    if settings.sheets_enabled:
        bg.add_task(sync_all)
    bg.add_task(notify_permission_event, permission_id, "RRHH_RECHAZADO", reason)
    return {"status": "ok"}


@app.post("/api/admin/email/test")
def test_email(user: dict = Depends(require_roles("RRHH"))):
    """Envía una prueba real a la dirección del administrador autenticado."""
    recipient = (user.get("email") or "").strip()
    if not recipient or recipient.endswith("@ersep.local"):
        raise HTTPException(
            status_code=422,
            detail="Configurá un email real para este usuario antes de probar las notificaciones.",
        )
    try:
        return send_test_email(recipient, user.get("nombre") or "Administrador")
    except Exception as exc:
        raise HTTPException(
            status_code=502,
            detail=f"No fue posible enviar el correo de prueba: {exc}",
        )


@app.get("/api/catalogos/oficinas")
def office_catalog(user: dict = Depends(get_current_user)):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT o.id,o.nombre,o.jefe_id,
                       trim(concat(j.nombre,' ',j.apellido)) jefe_nombre
                FROM oficinas o
                LEFT JOIN usuarios j ON j.id=o.jefe_id
                WHERE o.activo=TRUE
                ORDER BY o.nombre
            """)
            return {"items": cur.fetchall()}


@app.get("/api/admin/oficinas")
def list_offices(user: dict = Depends(require_roles("ADMIN"))):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT o.id,o.nombre,o.jefe_id,o.activo,o.created_at,o.updated_at,
                       trim(concat(j.nombre,' ',j.apellido)) jefe_nombre,j.email jefe_email,
                       COUNT(u.id) FILTER (WHERE u.activo=TRUE) agentes_activos
                FROM oficinas o
                LEFT JOIN usuarios j ON j.id=o.jefe_id
                LEFT JOIN usuarios u ON u.oficina_id=o.id
                GROUP BY o.id,j.nombre,j.apellido,j.email
                ORDER BY o.activo DESC,o.nombre
            """)
            return {"items": cur.fetchall()}


@app.post("/api/admin/oficinas")
def upsert_office(payload: AdminOfficeIn, user: dict = Depends(require_roles("ADMIN"))):
    name = payload.nombre.strip()
    with connection() as conn:
        with conn.cursor() as cur:
            if payload.jefe_id is not None:
                cur.execute("SELECT id,activo FROM usuarios WHERE id=%s", (payload.jefe_id,))
                boss = cur.fetchone()
                if not boss or not boss["activo"]:
                    raise HTTPException(status_code=422, detail="El jefe seleccionado debe ser un usuario activo.")
            if payload.id:
                cur.execute("SELECT id FROM oficinas WHERE id=%s", (payload.id,))
                if not cur.fetchone():
                    raise HTTPException(status_code=404, detail="Oficina inexistente.")
                cur.execute("SELECT id FROM oficinas WHERE lower(nombre)=lower(%s) AND id<>%s", (name, payload.id))
                if cur.fetchone():
                    raise HTTPException(status_code=409, detail="Ya existe otra Oficina con ese nombre.")
                cur.execute("""
                    UPDATE oficinas SET nombre=%s,jefe_id=%s,activo=%s,updated_at=NOW()
                    WHERE id=%s RETURNING id
                """, (name, payload.jefe_id, payload.activo, payload.id))
                office_id = cur.fetchone()["id"]
            else:
                cur.execute("""
                    INSERT INTO oficinas(nombre,jefe_id,activo)
                    VALUES (%s,%s,%s)
                    ON CONFLICT (nombre) DO UPDATE SET jefe_id=EXCLUDED.jefe_id,activo=EXCLUDED.activo,updated_at=NOW()
                    RETURNING id
                """, (name, payload.jefe_id, payload.activo))
                office_id = cur.fetchone()["id"]

            if payload.jefe_id is not None:
                cur.execute("""
                    INSERT INTO usuario_roles(usuario_id,rol_id)
                    SELECT %s,id FROM roles WHERE codigo='JEFE'
                    ON CONFLICT DO NOTHING
                """, (payload.jefe_id,))
            # Mantiene sincronizado el campo heredado sólo para integraciones antiguas.
            cur.execute("UPDATE usuarios SET area=%s,updated_at=NOW() WHERE oficina_id=%s", (name, office_id))
            conn.commit()
    return {"status": "ok", "id": office_id}


@app.get("/api/admin/usuarios")
def list_users(user: dict = Depends(require_roles("ADMIN"))):
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
              SELECT u.*,COALESCE(array_agg(DISTINCT r.codigo) FILTER (WHERE r.codigo IS NOT NULL),'{}') roles,
                     o.nombre oficina,
                     trim(concat(jefe.nombre,' ',jefe.apellido)) jefe_nombre,jefe.email jefe_email
              FROM usuarios u
              LEFT JOIN usuario_roles ur ON ur.usuario_id=u.id
              LEFT JOIN roles r ON r.id=ur.rol_id
              LEFT JOIN oficinas o ON o.id=u.oficina_id
              LEFT JOIN usuarios jefe ON jefe.id=o.jefe_id
              GROUP BY u.id,o.nombre,jefe.nombre,jefe.apellido,jefe.email
              ORDER BY u.activo DESC,u.apellido,u.nombre
            """)
            return {"items": [_serialize_user(r) for r in cur.fetchall()]}


@app.post("/api/admin/usuarios")
def upsert_user(payload: AdminUserIn, user: dict = Depends(require_roles("ADMIN"))):
    work_start = _parse_time(payload.jornada_desde)
    work_end = _parse_time(payload.jornada_hasta)
    if work_start >= work_end:
        raise HTTPException(status_code=422, detail="El fin de la jornada debe ser posterior al inicio.")

    username = payload.username.strip().lower()
    email = payload.email.lower()

    with connection() as conn:
        with conn.cursor() as cur:
            office_name = (payload.area or "").strip() or None
            office_id = payload.oficina_id
            if office_id is not None:
                cur.execute("SELECT id,nombre FROM oficinas WHERE id=%s AND activo=TRUE", (office_id,))
                office = cur.fetchone()
                if not office:
                    raise HTTPException(status_code=422, detail="La Oficina seleccionada no existe o está inactiva.")
                office_name = office["nombre"]

            cur.execute("SELECT * FROM usuarios WHERE email=%s", (email,))
            existing = cur.fetchone()
            cur.execute("SELECT id FROM usuarios WHERE lower(username)=lower(%s) AND email<>%s", (username, email))
            if cur.fetchone():
                raise HTTPException(status_code=409, detail="Ese nombre de usuario ya está en uso.")

            if existing:
                if payload.password:
                    cur.execute("""
                        UPDATE usuarios SET
                            username=%s,password_hash=%s,nombre=%s,apellido=%s,legajo=%s,dni=%s,
                            area=%s,oficina_id=%s,jornada_desde=%s,jornada_hasta=%s,activo=TRUE,updated_at=NOW()
                        WHERE id=%s RETURNING *
                    """, (
                        username,hash_password(payload.password),payload.nombre,payload.apellido,payload.legajo,payload.dni,
                        office_name,office_id,work_start,work_end,existing["id"],
                    ))
                else:
                    cur.execute("""
                        UPDATE usuarios SET
                            username=%s,nombre=%s,apellido=%s,legajo=%s,dni=%s,
                            area=%s,oficina_id=%s,jornada_desde=%s,jornada_hasta=%s,activo=TRUE,updated_at=NOW()
                        WHERE id=%s RETURNING *
                    """, (
                        username,payload.nombre,payload.apellido,payload.legajo,payload.dni,
                        office_name,office_id,work_start,work_end,existing["id"],
                    ))
                target = cur.fetchone()
            else:
                if not payload.password:
                    raise HTTPException(status_code=422, detail="Para crear un usuario nuevo debe indicar una clave inicial de al menos 6 caracteres.")
                cur.execute("""
                    INSERT INTO usuarios(
                        username,password_hash,email,nombre,apellido,legajo,dni,area,oficina_id,jornada_desde,jornada_hasta,activo
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE) RETURNING *
                """, (
                    username,hash_password(payload.password),email,payload.nombre,payload.apellido,payload.legajo,payload.dni,
                    office_name,office_id,work_start,work_end,
                ))
                target = cur.fetchone()

            cur.execute("DELETE FROM usuario_roles WHERE usuario_id=%s", (target["id"],))
            roles = set(payload.roles or ["AGENTE"])
            for role in roles:
                cur.execute("""
                    INSERT INTO usuario_roles(usuario_id,rol_id)
                    SELECT %s,id FROM roles WHERE codigo=%s ON CONFLICT DO NOTHING
                """, (target["id"], role))

            # Si la persona está definida como jefatura de alguna Oficina, el rol
            # JEFE es estructural y no puede perderse por una edición del usuario.
            cur.execute("""
                INSERT INTO usuario_roles(usuario_id,rol_id)
                SELECT %s,r.id FROM roles r
                WHERE r.codigo='JEFE'
                  AND EXISTS (SELECT 1 FROM oficinas o WHERE o.jefe_id=%s AND o.activo=TRUE)
                ON CONFLICT DO NOTHING
            """, (target["id"], target["id"]))

            # V6: no se crea ni actualiza una jefatura por agente. La jefatura
            # se resuelve exclusivamente desde la Oficina asignada.
            conn.commit()
            return {"status": "ok", "id": target["id"]}


@app.post("/api/admin/usuarios/{target_id}/estado")
def set_user_status(target_id: int, payload: AdminUserStatusIn, user: dict = Depends(require_roles("ADMIN"))):
    if target_id == user["id"] and not payload.activo:
        raise HTTPException(status_code=409, detail="No podés deshabilitar tu propio usuario.")
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM usuarios WHERE id=%s", (target_id,))
            target = cur.fetchone()
            if not target:
                raise HTTPException(status_code=404, detail="Usuario inexistente.")
            if (target.get("username") or "").lower() == settings.bootstrap_admin_username.lower() and not payload.activo:
                raise HTTPException(status_code=409, detail="La cuenta administradora inicial no puede deshabilitarse desde el panel.")
            cur.execute("UPDATE usuarios SET activo=%s,updated_at=NOW() WHERE id=%s", (payload.activo, target_id))
            if not payload.activo:
                cur.execute("UPDATE jefaturas SET fecha_hasta=CURRENT_DATE-1 WHERE (usuario_id=%s OR jefe_id=%s) AND fecha_hasta IS NULL", (target_id, target_id))
                cur.execute("UPDATE oficinas SET jefe_id=NULL,updated_at=NOW() WHERE jefe_id=%s", (target_id,))
            conn.commit()
    return {"status": "ok", "activo": payload.activo}


def _norm_header(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", text.strip().lower()).strip("_")


def _excel_bool(value, default=True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() in {"1", "si", "sí", "true", "verdadero", "activo", "x", "yes"}


def _excel_time(value, default: str) -> str:
    if value is None or str(value).strip() == "":
        return default
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    text = str(value).strip()
    match = re.match(r"^(\d{1,2}):(\d{2})", text)
    if not match:
        raise ValueError(f"Horario inválido: {text}")
    hh, mm = int(match.group(1)), int(match.group(2))
    if hh > 23 or mm > 59:
        raise ValueError(f"Horario inválido: {text}")
    return f"{hh:02d}:{mm:02d}"


def _parse_roles(value) -> list[str]:
    if value is None or str(value).strip() == "":
        return ["AGENTE"]
    parts = [x.strip().upper().replace("RR.HH.", "RRHH").replace("RRHH.", "RRHH") for x in re.split(r"[,;|]", str(value)) if x.strip()]
    allowed = {"AGENTE", "JEFE", "RRHH", "ADMIN"}
    invalid = [x for x in parts if x not in allowed]
    if invalid:
        raise ValueError(f"Roles inválidos: {', '.join(invalid)}")
    return sorted(set(parts)) or ["AGENTE"]


def _read_users_xlsx(content: bytes) -> list[dict]:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"No fue posible leer el archivo Excel: {exc}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="El Excel está vacío.")
    aliases = {
        "usuario": "username", "username": "username", "user": "username",
        "clave": "password", "password": "password", "contrasena": "password",
        "email": "email", "correo": "email",
        "nombre": "nombre", "apellido": "apellido", "legajo": "legajo", "dni": "dni",
        "oficina": "oficina", "area": "oficina",
        "jornada_desde": "jornada_desde", "desde": "jornada_desde",
        "jornada_hasta": "jornada_hasta", "hasta": "jornada_hasta",
        "roles": "roles", "rol": "roles", "activo": "activo",
    }
    headers = []
    for h in rows[0]:
        key = _norm_header(h)
        headers.append(aliases.get(key, key))
    required = {"username", "email", "nombre", "apellido", "legajo"}
    missing = sorted(required - set(headers))
    if missing:
        raise HTTPException(status_code=422, detail="Faltan columnas obligatorias: " + ", ".join(missing))
    result = []
    for excel_row, values in enumerate(rows[1:], start=2):
        raw = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        if not any(v not in (None, "") for v in raw.values()):
            continue
        result.append({"fila": excel_row, **raw})
    if not result:
        raise HTTPException(status_code=422, detail="No se encontraron filas de usuarios para procesar.")
    return result


def _validate_import_rows(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    valid, errors = [], []
    seen: dict[tuple[str, str], int] = {}
    with connection() as conn:
        with conn.cursor() as cur:
            for raw in rows:
                rownum = raw.get("fila")
                try:
                    username = str(raw.get("username") or "").strip().lower()
                    email = str(raw.get("email") or "").strip().lower()
                    legajo = str(raw.get("legajo") or "").strip()
                    password = str(raw.get("password") or "").strip() or None
                    roles = _parse_roles(raw.get("roles"))
                    jornada_desde = _excel_time(raw.get("jornada_desde"), "08:00")
                    jornada_hasta = _excel_time(raw.get("jornada_hasta"), "14:00")
                    office_name = str(raw.get("oficina") or "").strip() or None
                    item = AdminUserIn(
                        username=username, password=password, email=email,
                        nombre=str(raw.get("nombre") or "").strip(), apellido=str(raw.get("apellido") or "").strip(),
                        legajo=legajo, dni=str(raw.get("dni") or "").strip() or None, area=office_name,
                        oficina_id=None, jornada_desde=jornada_desde, jornada_hasta=jornada_hasta, roles=roles, jefe_email=None,
                    )
                    if _parse_time(jornada_desde) >= _parse_time(jornada_hasta):
                        raise ValueError("El fin de la jornada debe ser posterior al inicio.")

                    # Evita que un mismo Excel intente crear/actualizar dos veces la misma persona.
                    duplicate_refs = []
                    for kind, value in (("usuario", username), ("email", email), ("legajo", legajo)):
                        key = (kind, value.lower())
                        if key in seen:
                            duplicate_refs.append(f"{kind} repetido respecto de la fila {seen[key]}")
                        else:
                            seen[key] = rownum
                    if duplicate_refs:
                        raise ValueError("; ".join(duplicate_refs) + ".")

                    # Los tres identificadores deben resolver, si existen, a una única cuenta.
                    cur.execute("""
                        SELECT id,email,legajo,username
                        FROM usuarios
                        WHERE lower(email)=lower(%s) OR legajo=%s OR lower(username)=lower(%s)
                        ORDER BY id
                    """, (email, legajo, username))
                    matches = cur.fetchall()
                    distinct_ids = {m["id"] for m in matches}
                    if len(distinct_ids) > 1:
                        raise ValueError("Usuario, email y/o legajo pertenecen a cuentas distintas en la base actual.")
                    existing = matches[0] if matches else None
                    if not existing and not password:
                        raise ValueError("Usuario nuevo sin clave inicial (mínimo 6 caracteres).")
                    valid.append({
                        "fila": rownum, "payload": item.model_dump(), "oficina": office_name,
                        "activo": _excel_bool(raw.get("activo"), True), "existente_id": existing["id"] if existing else None,
                    })
                except (ValidationError, ValueError, HTTPException) as exc:
                    msg = exc.detail if isinstance(exc, HTTPException) else str(exc)
                    errors.append({"fila": rownum, "error": msg})
    return valid, errors


def _apply_import_item(cur, item: dict):
    data = item["payload"]
    office_id = None
    office_name = item.get("oficina")
    if office_name:
        cur.execute("SELECT id FROM oficinas WHERE lower(trim(nombre))=lower(trim(%s)) ORDER BY activo DESC,id LIMIT 1", (office_name,))
        office = cur.fetchone()
        if office:
            office_id = office["id"]
            cur.execute("UPDATE oficinas SET activo=TRUE,updated_at=NOW() WHERE id=%s", (office_id,))
        else:
            cur.execute("INSERT INTO oficinas(nombre,activo) VALUES (%s,TRUE) RETURNING id", (office_name,))
            office_id = cur.fetchone()["id"]
    work_start = _parse_time(data["jornada_desde"]); work_end = _parse_time(data["jornada_hasta"])
    existing_id = item.get("existente_id")
    if existing_id:
        if data.get("password"):
            cur.execute("""
                UPDATE usuarios SET username=%s,password_hash=%s,email=%s,nombre=%s,apellido=%s,legajo=%s,dni=%s,
                    area=%s,oficina_id=%s,jornada_desde=%s,jornada_hasta=%s,activo=%s,updated_at=NOW()
                WHERE id=%s RETURNING id
            """, (data["username"],hash_password(data["password"]),data["email"],data["nombre"],data["apellido"],data["legajo"],data.get("dni"),office_name,office_id,work_start,work_end,item["activo"],existing_id))
        else:
            cur.execute("""
                UPDATE usuarios SET username=%s,email=%s,nombre=%s,apellido=%s,legajo=%s,dni=%s,
                    area=%s,oficina_id=%s,jornada_desde=%s,jornada_hasta=%s,activo=%s,updated_at=NOW()
                WHERE id=%s RETURNING id
            """, (data["username"],data["email"],data["nombre"],data["apellido"],data["legajo"],data.get("dni"),office_name,office_id,work_start,work_end,item["activo"],existing_id))
        user_id = cur.fetchone()["id"]
    else:
        cur.execute("""
            INSERT INTO usuarios(username,password_hash,email,nombre,apellido,legajo,dni,area,oficina_id,jornada_desde,jornada_hasta,activo)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (data["username"],hash_password(data["password"]),data["email"],data["nombre"],data["apellido"],data["legajo"],data.get("dni"),office_name,office_id,work_start,work_end,item["activo"]))
        user_id = cur.fetchone()["id"]
    cur.execute("DELETE FROM usuario_roles WHERE usuario_id=%s", (user_id,))
    for role in data.get("roles") or ["AGENTE"]:
        cur.execute("INSERT INTO usuario_roles(usuario_id,rol_id) SELECT %s,id FROM roles WHERE codigo=%s ON CONFLICT DO NOTHING", (user_id, role))
    # La condición de jefatura pertenece a la Oficina: una importación no debe quitar ese acceso.
    cur.execute("""
        INSERT INTO usuario_roles(usuario_id,rol_id)
        SELECT %s,r.id FROM roles r
        WHERE r.codigo='JEFE' AND EXISTS (SELECT 1 FROM oficinas o WHERE o.jefe_id=%s AND o.activo=TRUE)
        ON CONFLICT DO NOTHING
    """, (user_id, user_id))
    return user_id


@app.post("/api/admin/importar-usuarios/validar")
async def validate_users_excel(file: UploadFile = File(...), user: dict = Depends(require_roles("ADMIN"))):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="El archivo debe estar en formato .xlsx.")
    rows = _read_users_xlsx(await file.read())
    valid, errors = _validate_import_rows(rows)
    return {
        "status": "ok" if not errors else "con_errores", "total": len(rows),
        "validos": len(valid), "errores": len(errors), "detalle_errores": errors[:100],
        "puede_importar": bool(valid) and not errors,
    }


@app.post("/api/admin/importar-usuarios/aplicar")
async def apply_users_excel(file: UploadFile = File(...), user: dict = Depends(require_roles("ADMIN"))):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="El archivo debe estar en formato .xlsx.")
    rows = _read_users_xlsx(await file.read())
    valid, errors = _validate_import_rows(rows)
    if errors:
        raise HTTPException(status_code=422, detail=f"El Excel tiene {len(errors)} fila(s) con errores. Validalo nuevamente antes de importar.")
    with connection() as conn:
        with conn.cursor() as cur:
            for item in valid:
                _apply_import_item(cur, item)
        conn.commit()
    return {"status": "ok", "importados": len(valid), "message": f"Se procesaron {len(valid)} usuarios correctamente."}


@app.post("/api/admin/limpieza/permisos")
def clear_permissions(payload: CleanupIn, user: dict = Depends(require_roles("ADMIN"))):
    if payload.confirmacion.strip().upper() != "LIMPIAR PERMISOS":
        raise HTTPException(status_code=422, detail="Confirmación inválida. Escribí LIMPIAR PERMISOS.")
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) cantidad FROM permisos_salida")
            count = cur.fetchone()["cantidad"]
            cur.execute("DELETE FROM permisos_salida")
        conn.commit()
    return {"status": "ok", "eliminados": count, "message": f"Se eliminaron {count} permisos y sus movimientos asociados."}


@app.post("/api/admin/limpieza/maestros")
def reset_master_data(payload: CleanupIn, user: dict = Depends(require_roles("ADMIN"))):
    if payload.confirmacion.strip().upper() != "REINICIAR MAESTROS":
        raise HTTPException(status_code=422, detail="Confirmación inválida. Escribí REINICIAR MAESTROS.")
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM usuarios WHERE id=%s OR lower(username)=lower(%s)", (user["id"], settings.bootstrap_admin_username))
            preserve = [r["id"] for r in cur.fetchall()] or [user["id"]]
            cur.execute("DELETE FROM permisos_salida")
            cur.execute("DELETE FROM jefaturas")
            cur.execute("UPDATE oficinas SET jefe_id=NULL")
            cur.execute("UPDATE usuarios SET oficina_id=NULL,area=NULL WHERE id=ANY(%s)", (preserve,))
            cur.execute("DELETE FROM oficinas")
            cur.execute("DELETE FROM google_sheets_oauth_states WHERE usuario_id <> ALL(%s)", (preserve,))
            cur.execute("DELETE FROM usuarios WHERE id <> ALL(%s)", (preserve,))
            cur.execute("DELETE FROM usuario_roles WHERE usuario_id=ANY(%s)", (preserve,))
            for preserved_id in preserve:
                cur.execute("INSERT INTO usuario_roles(usuario_id,rol_id) SELECT %s,id FROM roles WHERE codigo='ADMIN' ON CONFLICT DO NOTHING", (preserved_id,))
                cur.execute("INSERT INTO usuario_roles(usuario_id,rol_id) SELECT %s,id FROM roles WHERE codigo='RRHH' ON CONFLICT DO NOTHING", (preserved_id,))
                cur.execute("UPDATE usuarios SET activo=TRUE,updated_at=NOW() WHERE id=%s", (preserved_id,))
        conn.commit()
    return {"status": "ok", "message": "Maestros reiniciados. Se conservaron las cuentas administradoras protegidas."}


@app.get("/api/sheets/status")
def sheets_status(user: dict = Depends(require_roles("RRHH"))):
    return integration_status()


@app.post("/api/sheets/connect")
def sheets_connect(user: dict = Depends(require_roles("RRHH"))):
    try:
        return {"authorization_url": create_authorization_url(user)}
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=f"No fue posible iniciar la autorización de Google Sheets: {exc}",
        )


@app.get("/api/google-sheets/callback")
def google_sheets_callback(
    state: str | None = Query(default=None),
    code: str | None = Query(default=None),
    error: str | None = Query(default=None),
):
    base = settings.permisos_frontend_url.rstrip("/") + "/"
    if error:
        return RedirectResponse(
            url=f"{base}?sheets=error&message={error}",
            status_code=302,
        )
    if not state or not code:
        return RedirectResponse(
            url=f"{base}?sheets=error&message=respuesta_incompleta",
            status_code=302,
        )

    try:
        finish_authorization(state, code)
        return RedirectResponse(
            url=f"{base}?sheets=connected",
            status_code=302,
        )
    except Exception as exc:
        from urllib.parse import quote
        return RedirectResponse(
            url=f"{base}?sheets=error&message={quote(str(exc))}",
            status_code=302,
        )


@app.post("/api/sheets/disconnect")
def sheets_disconnect(user: dict = Depends(require_roles("RRHH"))):
    try:
        return disconnect_sheets()
    except Exception as exc:
        raise HTTPException(
            status_code=502,
            detail=f"No fue posible desconectar Google Sheets: {exc}",
        )


@app.post("/api/sheets/sync")
def sync_sheets(user: dict = Depends(require_roles("RRHH"))):
    try:
        return sync_all()
    except Exception as exc:
        raise HTTPException(
            status_code=502,
            detail=f"No fue posible sincronizar Google Sheets: {exc}",
        )
